import time
from pathlib import Path
from openpyxl import load_workbook
from typing import TYPE_CHECKING
from odoo.modules.module import get_module_path

if TYPE_CHECKING:
    from odoo.orm.environments import Environment

    env: Environment = {}


emails_path = Path(get_module_path("pkf_nominas")) / "scripts/emails.xlsx"

if not emails_path.exists():
    raise ValueError(f"El archivo {str(emails_path)} no existe.")


wb = load_workbook(str(emails_path))
ws = wb["MS365"]

rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

data = [{k: row[i] for i, k in enumerate(headers)} for row in rows[1:]]


for item in data:
    code = item["Codigo"]
    email = item["Correo"]

    emp = env["hr.employee"].search([("ev_codigo", "=", code)], limit=1)

    if not emp:
        print("No se encontro el empleado")
        continue

    old_email = emp.email
    emp.invalidate_recordset()

    if old_email == email:
        continue

    e = emp.write({"email": email, "work_email": email})
    if e:
        env.cr.commit()
        print(f"{emp.name} ha sido actualizado de {emp.email} a {email}")
    time.sleep(0.1)
